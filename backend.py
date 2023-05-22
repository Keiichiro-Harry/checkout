from flask import Flask, request, jsonify, render_template
import stripe
import smtplib
from email.mime.text import MIMEText
import openpyxl
import requests
from openpyxl import load_workbook

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/submit', methods=['POST'])
def submit_form():
    return handle_form_submission()


def handle_form_submission():
    data = request.get_json()

    email = data['email']
    firstName = data['firstName']
    lastName = data['lastName']
    phoneticFirstName = data['phoneticFirstName']
    phoneticLastName = data['phoneticLastName']
    phone = data['phone']
    courses = data['courses']

    appli = [False, False, False]

    if 'A' in courses:
        appli[0] = True
    if 'B' in courses:
        appli[1] = True
    if 'C' in courses:
        appli[2] = True

    fee = [2000, 3000, 4000]
    totalAmount = 0
    appli_name = ["講座A", "講座B", "講座C"]
    appli_str = ""

    for i in range(len(fee)):
        if appli[i]:
            totalAmount += fee[i]
            if appli_str != "":
                appli_str += ", "
            appli_str += appli_name[i]

    #Stripeでの決済フォームの作成
    url_to_form = create_payment_form(totalAmount, "JPY", appli_str)

    # 確認メールの作成と送信
    subject = "お申込みの確認とお支払いフォーム"
    message = f"こんにちは、{firstName} {lastName}様。\n\nお申込みありがとうございます。回答内容は以下の通りです。\n\n\nお名前 : {firstName} {lastName} 様\n\nメールアドレス : {email}\n\nお電話番号 : {phone}\n\nお申込み講座 : {appli_str}\n\nお支払料金 : {totalAmount}円\n\n\n以上が正しければ、以下の決済フォームから、お支払いをお願いいたします。\n\n{url_to_form}\n\n\nこの度はお申込みいただきありがとうございました。\n\n\n担当者"
    send_email(subject, message, email)

    # データをエクセルにまとめる
    write_to_excel(data)

    # フロントエンドへのデータの受け渡し
    response_data = {
        'message': 'データを受け取りました。処理しました。',
        # 必要な情報を追加する
        # ...
    }

    return jsonify(response_data)

def create_payment_form(amount, currency, description): #参照：https://stripe.com/docs/payments/checkout/migrating-prices
    print("in create payment form")
    stripe.api_key = ''
    session = stripe.checkout.Session.create(
        payment_method_types=['card'],
        line_items=[{
            'price_data': {
                'currency': currency,
                'unit_amount': amount,
                'product_data': {
                    'name': 'Courses',
                    'description': description,
                    #'images': ['https://example.com/t-shirt.png'],
                },
            },
            'quantity': 1,
        }],
        mode = 'payment',
        success_url = '',
        cancel_url = ''
    )

    #print(session)  # レスポンスデータを出力して確認する

    return session.url


def send_email(subject, message, recipient):
    sender = ""  # 送信元メールアドレス
    password = ""  # 送信元メールアドレスのパスワード

    # メールの構築
    msg = MIMEText(message)
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = recipient

    # SMTPサーバに接続してメール送信
    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(sender, password)
        server.send_message(msg)

# データをエクセルにまとめる関数
def write_to_excel(data):
    # 既存のExcelファイルを読み込む
    workbook = load_workbook('file.xlsx')
    sheet = workbook.active

    # データをExcelファイルに書き込む
    row = sheet.max_row + 1  # 新しいデータを書き込む行番号
    sheet.cell(row=row, column=1).value = data['email']
    sheet.cell(row=row, column=2).value = data['firstName']
    sheet.cell(row=row, column=3).value = data['lastName']
    sheet.cell(row=row, column=4).value = data['phoneticFirstName']
    sheet.cell(row=row, column=5).value = data['phoneticLastName']
    sheet.cell(row=row, column=6).value = data['phone']
    sheet.cell(row=row, column=7).value = 'A' in data['courses']
    sheet.cell(row=row, column=8).value = 'B' in data['courses']
    sheet.cell(row=row, column=9).value = 'C' in data['courses']

    # Excelファイルを保存
    workbook.save('file.xlsx')

# メイン処理
def main():
    #データの受信/メール送信/フロントエンドへの返信/Excelへの保存
    handle_form_submission()

if __name__ == '__main__':
    app.run()
    print("Flask app is running!")
    main()
